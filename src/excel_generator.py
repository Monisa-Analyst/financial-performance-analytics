import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_excel(excel_path="financial_analysis.xlsx"):
    print(f"Compiling styled reports in Excel workbook: {excel_path}")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Source Excel database not found at {excel_path}. Run db_init.py first.")
        
    wb = openpyxl.load_workbook(excel_path)
    
    # Remove existing styled sheets if they exist to prevent duplicates
    for sheet in ["Executive Summary", "Revenue & Margin Analysis", "Budget vs Actual Report", "Segment Profitability"]:
        if sheet in wb.sheetnames:
            del wb[sheet]
            
    # Load raw data lists from fact/dimension sheets to know the rows we need to generate formulas for
    ws_trans = wb["fact_transactions"]
    ws_depts = wb["dim_departments"]
    ws_regions = wb["dim_regions"]
    ws_products = wb["dim_products"]
    
    # Extract unique sorted months from transactions
    dates = []
    for row in ws_trans.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0]:
            dates.append(row[0][:7]) # YYYY-MM
    months = sorted(list(set(dates)))
    
    # Extract department names
    departments = []
    for row in ws_depts.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0]:
            departments.append(row[0])
            
    # Extract region names
    regions = []
    for row in ws_regions.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0]:
            regions.append(row[0])
            
    # Extract product lines
    products = []
    for row in ws_products.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0]:
            products.append(row[0])
            
    # Styling configurations
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    section_font = Font(name=font_family, size=12, bold=True, color="2C3E50")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    summary_fill = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='7F8C8D'),
        bottom=Side(style='double', color='2C3E50')
    )

    # --- SHEET 1: REVENUE & MARGIN ANALYSIS ---
    ws_rev = wb.create_sheet(title="Revenue & Margin Analysis", index=0)
    ws_rev.views.sheetView[0].showGridLines = True
    ws_rev.cell(row=1, column=1, value="Monthly Revenue & Profitability Analysis").font = title_font
    
    rev_headers = ["Month", "Revenue", "Expenses", "Net Profit", "Profit Margin", "Running Total Revenue"]
    for c_idx, h in enumerate(rev_headers, start=1):
        cell = ws_rev.cell(row=2, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for r_idx, month in enumerate(months):
        excel_row = r_idx + 3
        # Add month name
        ws_rev.cell(row=excel_row, column=1, value=month).font = normal_font
        
        # SUMIFS formulas querying fact_transactions
        # Revenue is Column C, Date is Column B
        ws_rev.cell(row=excel_row, column=2, value=f'=SUMIFS(fact_transactions!C:C, fact_transactions!B:B, A{excel_row}&"-*")')
        # Expenses is Column D
        ws_rev.cell(row=excel_row, column=3, value=f'=SUMIFS(fact_transactions!D:D, fact_transactions!B:B, A{excel_row}&"-*")')
        
        # Net Profit = Revenue - Expenses
        ws_rev.cell(row=excel_row, column=4, value=f"=B{excel_row}-C{excel_row}")
        # Profit Margin = Net Profit / Revenue
        ws_rev.cell(row=excel_row, column=5, value=f"=IF(B{excel_row}=0, 0, D{excel_row}/B{excel_row})")
        # Running Total = SUM(B$3:B{current})
        ws_rev.cell(row=excel_row, column=6, value=f"=SUM(B$3:B{excel_row})")
        
        # Style cells
        for c_idx in range(1, 7):
            cell = ws_rev.cell(row=excel_row, column=c_idx)
            cell.font = normal_font
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
                
            if c_idx in [2, 3, 4, 6]:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 5:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")
                
    # Add Totals Row
    tot_row = len(months) + 3
    ws_rev.cell(row=tot_row, column=1, value="Total").font = bold_font
    ws_rev.cell(row=tot_row, column=2, value=f"=SUM(B3:B{tot_row-1})")
    ws_rev.cell(row=tot_row, column=3, value=f"=SUM(C3:C{tot_row-1})")
    ws_rev.cell(row=tot_row, column=4, value=f"=SUM(D3:D{tot_row-1})")
    ws_rev.cell(row=tot_row, column=5, value=f"=IF(B{tot_row}=0, 0, D{tot_row}/B{tot_row})")
    ws_rev.cell(row=tot_row, column=6, value="") # running total blank
    
    for c_idx in range(1, 7):
        cell = ws_rev.cell(row=tot_row, column=c_idx)
        cell.font = bold_font
        cell.fill = summary_fill
        cell.border = double_bottom_border
        if c_idx in [2, 3, 4]:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        elif c_idx == 5:
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")

    # --- SHEET 2: BUDGET VS ACTUAL REPORT ---
    ws_bva = wb.create_sheet(title="Budget vs Actual Report", index=1)
    ws_bva.views.sheetView[0].showGridLines = True
    ws_bva.cell(row=1, column=1, value="Budget vs Actual Variance Report (by Department)").font = title_font
    
    bva_headers = ["Department", "Actual Revenue", "Budget Revenue", "Revenue Variance", "Actual Expenses", "Budget Expenses", "Expense Variance"]
    for c_idx, h in enumerate(bva_headers, start=1):
        cell = ws_bva.cell(row=2, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for r_idx, dept in enumerate(departments):
        excel_row = r_idx + 3
        ws_bva.cell(row=excel_row, column=1, value=dept).font = bold_font
        
        # SUMIFS with XLOOKUP to map department name to ID
        # Actual Revenue
        ws_bva.cell(row=excel_row, column=2, value=f'=SUMIFS(fact_transactions!C:C, fact_transactions!E:E, XLOOKUP(A{excel_row}, dim_departments!B:B, dim_departments!A:A))')
        # Budget Revenue (fact_budgets column E is budgeted_revenue, column C is dept_id)
        ws_bva.cell(row=excel_row, column=3, value=f'=SUMIFS(fact_budgets!E:E, fact_budgets!C:C, XLOOKUP(A{excel_row}, dim_departments!B:B, dim_departments!A:A))')
        # Revenue Variance = Actual - Budget
        ws_bva.cell(row=excel_row, column=4, value=f"=B{excel_row}-C{excel_row}")
        
        # Actual Expenses
        ws_bva.cell(row=excel_row, column=5, value=f'=SUMIFS(fact_transactions!D:D, fact_transactions!E:E, XLOOKUP(A{excel_row}, dim_departments!B:B, dim_departments!A:A))')
        # Budget Expenses (fact_budgets column F is budgeted_expenses)
        ws_bva.cell(row=excel_row, column=6, value=f'=SUMIFS(fact_budgets!F:F, fact_budgets!C:C, XLOOKUP(A{excel_row}, dim_departments!B:B, dim_departments!A:A))')
        # Expense Variance = Budget - Actual (Positive is savings)
        ws_bva.cell(row=excel_row, column=7, value=f"=F{excel_row}-E{excel_row}")
        
        for c_idx in range(1, 8):
            cell = ws_bva.cell(row=excel_row, column=c_idx)
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
                
            if c_idx > 1:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
                cell.font = normal_font
                
    # Totals Row
    tot_row = len(departments) + 3
    ws_bva.cell(row=tot_row, column=1, value="Total").font = bold_font
    ws_bva.cell(row=tot_row, column=2, value=f"=SUM(B3:B{tot_row-1})")
    ws_bva.cell(row=tot_row, column=3, value=f"=SUM(C3:C{tot_row-1})")
    ws_bva.cell(row=tot_row, column=4, value=f"=B{tot_row}-C{tot_row}")
    ws_bva.cell(row=tot_row, column=5, value=f"=SUM(E3:E{tot_row-1})")
    ws_bva.cell(row=tot_row, column=6, value=f"=SUM(F3:F{tot_row-1})")
    ws_bva.cell(row=tot_row, column=7, value=f"=F{tot_row}-E{tot_row}")
    
    for c_idx in range(1, 8):
        cell = ws_bva.cell(row=tot_row, column=c_idx)
        cell.font = bold_font
        cell.fill = summary_fill
        cell.border = double_bottom_border
        if c_idx > 1:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    # --- SHEET 3: SEGMENT PROFITABILITY ---
    ws_seg = wb.create_sheet(title="Segment Profitability", index=2)
    ws_seg.views.sheetView[0].showGridLines = True
    
    # Regional table
    ws_seg.cell(row=1, column=1, value="Profitability by Region").font = section_font
    reg_headers = ["Region", "Revenue", "Expenses", "Net Profit", "Margin"]
    for c_idx, h in enumerate(reg_headers, start=1):
        cell = ws_seg.cell(row=2, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for r_idx, region in enumerate(regions):
        excel_row = r_idx + 3
        ws_seg.cell(row=excel_row, column=1, value=region).font = bold_font
        # SUMIFS with XLOOKUP to map region name to ID (fact_transactions Column F is region_id)
        ws_seg.cell(row=excel_row, column=2, value=f'=SUMIFS(fact_transactions!C:C, fact_transactions!F:F, XLOOKUP(A{excel_row}, dim_regions!B:B, dim_regions!A:A))')
        ws_seg.cell(row=excel_row, column=3, value=f'=SUMIFS(fact_transactions!D:D, fact_transactions!F:F, XLOOKUP(A{excel_row}, dim_regions!B:B, dim_regions!A:A))')
        ws_seg.cell(row=excel_row, column=4, value=f"=B{excel_row}-C{excel_row}")
        ws_seg.cell(row=excel_row, column=5, value=f"=IF(B{excel_row}=0,0,D{excel_row}/B{excel_row})")
        
        for c_idx in range(1, 6):
            cell = ws_seg.cell(row=excel_row, column=c_idx)
            cell.border = thin_border
            if c_idx in [2, 3, 4]:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
                cell.font = normal_font
            elif c_idx == 5:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
                cell.font = normal_font
                
    # Regional Total
    reg_tot = len(regions) + 3
    ws_seg.cell(row=reg_tot, column=1, value="Total").font = bold_font
    ws_seg.cell(row=reg_tot, column=2, value=f"=SUM(B3:B{reg_tot-1})")
    ws_seg.cell(row=reg_tot, column=3, value=f"=SUM(C3:C{reg_tot-1})")
    ws_seg.cell(row=reg_tot, column=4, value=f"=B{reg_tot}-C{reg_tot}")
    ws_seg.cell(row=reg_tot, column=5, value=f"=IF(B{reg_tot}=0,0,D{reg_tot}/B{reg_tot})")
    
    for c_idx in range(1, 6):
        cell = ws_seg.cell(row=reg_tot, column=c_idx)
        cell.font = bold_font
        cell.fill = summary_fill
        cell.border = double_bottom_border
        if c_idx in [2, 3, 4]:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        elif c_idx == 5:
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")
            
    # Product Line table (offset by 2 columns)
    ws_seg.cell(row=1, column=7, value="Profitability by Product Line").font = section_font
    prod_headers = ["Product Line", "Revenue", "Expenses", "Net Profit", "Margin"]
    for c_idx, h in enumerate(prod_headers, start=7):
        cell = ws_seg.cell(row=2, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for r_idx, product in enumerate(products):
        excel_row = r_idx + 3
        ws_seg.cell(row=excel_row, column=7, value=product).font = bold_font
        # SUMIFS with XLOOKUP to map product name to ID (fact_transactions Column G is product_id)
        ws_seg.cell(row=excel_row, column=8, value=f'=SUMIFS(fact_transactions!C:C, fact_transactions!G:G, XLOOKUP(G{excel_row}, dim_products!B:B, dim_products!A:A))')
        ws_seg.cell(row=excel_row, column=9, value=f'=SUMIFS(fact_transactions!D:D, fact_transactions!G:G, XLOOKUP(G{excel_row}, dim_products!B:B, dim_products!A:A))')
        ws_seg.cell(row=excel_row, column=10, value=f"=H{excel_row}-I{excel_row}")
        ws_seg.cell(row=excel_row, column=11, value=f"=IF(H{excel_row}=0,0,J{excel_row}/H{excel_row})")
        
        for c_idx in range(7, 12):
            cell = ws_seg.cell(row=excel_row, column=c_idx)
            cell.border = thin_border
            if c_idx in [8, 9, 10]:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
                cell.font = normal_font
            elif c_idx == 11:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
                cell.font = normal_font
                
    # Product Total
    prod_tot = len(products) + 3
    ws_seg.cell(row=prod_tot, column=7, value="Total").font = bold_font
    ws_seg.cell(row=prod_tot, column=8, value=f"=SUM(H3:H{prod_tot-1})")
    ws_seg.cell(row=prod_tot, column=9, value=f"=SUM(I3:I{prod_tot-1})")
    ws_seg.cell(row=prod_tot, column=10, value=f"=H{prod_tot}-I{prod_tot}")
    ws_seg.cell(row=prod_tot, column=11, value=f"=IF(H{prod_tot}=0,0,J{prod_tot}/H{prod_tot})")
    
    for c_idx in range(7, 12):
        cell = ws_seg.cell(row=prod_tot, column=c_idx)
        cell.font = bold_font
        cell.fill = summary_fill
        cell.border = double_bottom_border
        if c_idx in [8, 9, 10]:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        elif c_idx == 11:
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")

    # --- SHEET 4: EXECUTIVE SUMMARY ---
    ws_exec = wb.create_sheet(title="Executive Summary", index=0) # Make it the very first sheet
    ws_exec.views.sheetView[0].showGridLines = True
    ws_exec.cell(row=2, column=2, value="FinSight Financial Performance Summary").font = title_font
    
    # KPI card compiler
    def create_kpi_card(ws, start_col, start_row, label, value_formula, number_format):
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+1)
        ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+1)
        
        lbl_cell = ws.cell(row=start_row, column=start_col, value=label)
        lbl_cell.font = Font(name=font_family, size=9, bold=True, color="7F8C8D")
        lbl_cell.alignment = Alignment(horizontal="center")
        
        val_cell = ws.cell(row=start_row+1, column=start_col, value=value_formula)
        val_cell.font = Font(name=font_family, size=14, bold=True, color="1F4E78")
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.number_format = number_format
        
        for r in range(start_row, start_row+2):
            for c in range(start_col, start_col+2):
                ws.cell(row=r, column=c).border = thin_border
                
    # Add KPI Cards pointing to the totals row of Revenue sheet
    create_kpi_card(ws_exec, 2, 4, "TOTAL REVENUE", f"='Revenue & Margin Analysis'!B{tot_row}", "$#,##0.00")
    create_kpi_card(ws_exec, 5, 4, "TOTAL NET PROFIT", f"='Revenue & Margin Analysis'!D{tot_row}", "$#,##0.00")
    create_kpi_card(ws_exec, 8, 4, "NET PROFIT MARGIN", f"='Revenue & Margin Analysis'!E{tot_row}", "0.0%")
    
    # Overview metrics table
    ws_exec.cell(row=8, column=2, value="Corporate Performance Overview").font = section_font
    overview_headers = ["Key Metrics", "Current Year Value", "Target Budget", "Variance"]
    for c_idx, h in enumerate(overview_headers, start=2):
        cell = ws_exec.cell(row=9, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    metrics_list = [
        ("Total Sales Revenue", f"='Revenue & Margin Analysis'!B{tot_row}", f"='Budget vs Actual Report'!C{len(departments)+3}", "=C10-D10"),
        ("Operational Expenses", f"='Revenue & Margin Analysis'!C{tot_row}", f"='Budget vs Actual Report'!F{len(departments)+3}", "=D11-C11"),
        ("Net Operating Profit", f"='Revenue & Margin Analysis'!D{tot_row}", "=C10-C11", "=C12-D12"),
    ]
    
    for r_idx, (m_name, act_f, bud_f, var_f) in enumerate(metrics_list, start=10):
        c1 = ws_exec.cell(row=r_idx, column=2, value=m_name)
        c2 = ws_exec.cell(row=r_idx, column=3, value=act_f)
        c3 = ws_exec.cell(row=r_idx, column=4, value=bud_f)
        c4 = ws_exec.cell(row=r_idx, column=5, value=var_f)
        
        c1.font = bold_font
        c1.border = thin_border
        
        for c in [c2, c3, c4]:
            c.font = normal_font
            c.border = thin_border
            c.number_format = "$#,##0.00"
            c.alignment = Alignment(horizontal="right")
            
    for c in range(2, 6):
        ws_exec.cell(row=13, column=c).border = Border(bottom=Side(style='double', color='1F4E78'))
        
    # Auto-fit column widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if val.startswith('='):
                    val = " $1,000,000.00 "
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    wb.save(excel_path)
    print("Styled report sheets compiled successfully inside the Excel workbook!")

if __name__ == "__main__":
    out_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "financial_analysis.xlsx")
    create_styled_excel(out_file)
